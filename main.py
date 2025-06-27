import shutil
import threading
import time
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from itertools import chain

class Image2ExcelCore:
    def __init__(self):
        self._pause_event = threading.Event()
        self._stop_event = threading.Event()
        self._thread = None
        self.log_queue = None
        self.progress_queue = None

    def _log(self, code, msg, tag):
        if self.log_queue:
            self.log_queue((code, msg, tag))

    def _progress(self, total, current):
        if self.progress_queue:
            self.progress_queue((total, current))

    def start(self, product_file_path, image_folder, matched_folder, suffixes, log_queue, progress_queue):
        self.product_file = Path(product_file_path)
        self.image_folder = Path(image_folder)
        base_matched = Path(matched_folder)
        if not base_matched.exists():
            self._log(None, f"❌ Lỗi: Thư mục {base_matched} không tồn tại.", 'error')
            return
        if base_matched.resolve() == self.image_folder.resolve():
            self._log(None, "❌ Lỗi: Thư mục lưu không được trùng với thư mục ảnh gốc.", 'error')
            return
        self.matched_folder = base_matched / "Image2Excel_Export"
        self.suffixes = suffixes
        self.log_queue = log_queue
        self.progress_queue = progress_queue
        self._pause_event.set()
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._run)
        self._thread.start()

    def _run(self):
        try:
            if not self.product_file.exists():
                self._log(None, f"❌ Lỗi: Không tìm thấy file {self.product_file}.", 'error')
                return

            if self.product_file.suffix.lower() == '.txt':
                with open(self.product_file, 'r', encoding='utf-8') as f:
                    codes = [line.strip() for line in f if line.strip()]
            elif self.product_file.suffix.lower() == '.xlsx':
                from openpyxl import load_workbook
                wb = load_workbook(self.product_file, data_only=True)
                sheet = wb.active
                codes = [str(cell.value).strip() for cell in sheet['A'] if cell.value]
            else:
                self._log(None, "❌ Định dạng file không hợp lệ.", 'error')
                return

            if self.matched_folder.exists():
                for file in chain(
                    self.matched_folder.glob("*.[jJ][pP][gG]"),
                    self.matched_folder.glob("*.[pP][nN][gG]"),
                    self.matched_folder.glob("*.[jJ][pP][eE][gG]")
                ):
                    file.unlink()
                self._log(None, f"🗑️ Đã xóa file ảnh trong {self.matched_folder}", 'info')
            self.matched_folder.mkdir(exist_ok=True)

            wb = Workbook()
            ws = wb.active
            headers = ["Mã SP"] + [f"Ảnh {suffix}" for suffix in self.suffixes]
            ws.append(headers)
            ws.column_dimensions['A'].width = 20
            for col in range(2, len(self.suffixes) + 2):
                ws.column_dimensions[chr(64 + col)].width = 35

            self.files_by_suffix = {suffix: [] for suffix in self.suffixes}
            for file in self.image_folder.rglob("*"):
                if file.suffix.lower() in ['.jpg', '.png', '.jpeg']:
                    if self.matched_folder in file.parents:
                        continue
                    name = file.stem.lower()
                    normalized_name = name.replace("-", "").replace("_", "").replace(".", "").replace(" ", "")
                    for suffix in self.suffixes:
                        if normalized_name.endswith(suffix.lower()):
                            self.files_by_suffix[suffix].append((file, normalized_name))

            total_codes = len(codes)
            for idx, code in enumerate(codes, start=2):
                if self._stop_event.is_set():
                    self._log(code, "🛑 Đã dừng xử lý.", 'error')
                    break
                while not self._pause_event.is_set():
                    time.sleep(0.1)

                cell = ws.cell(row=idx, column=1)
                cell.value = code
                cell.alignment = Alignment(horizontal="center", vertical="center")

                found_images = []
                for col, suffix in enumerate(self.suffixes, start=2):
                    img_path = self._find_image(code, suffix)
                    if img_path:
                        copied = shutil.copy(img_path, self.matched_folder)
                        excel_img = ExcelImage(copied)
                        excel_img.width, excel_img.height = 200, 200
                        ws.add_image(excel_img, f"{chr(64 + col)}{idx}")
                        found_images.append(suffix)

                ws.row_dimensions[idx].height = 150

                if len(found_images) == len(self.suffixes):
                    self._log(code, f"✅ {code}: OK", 'ok')
                else:
                    missing = [s for s in self.suffixes if s not in found_images]
                    self._log(code, f"⚠️ {code}: Thiếu ảnh {', '.join(missing)}", 'warning')

                self._progress(total_codes, idx - 1)

            out_file = self.matched_folder / f"output_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
            wb.save(out_file)
            self._log(None, f"📦 Đã lưu Excel: {out_file}", 'info')

        except Exception as e:
            self._log(None, f"❌ Lỗi: {e}", 'error')

    def _find_image(self, product_code, suffix):
        code = product_code.lower()
        normalized_code = code.replace("-", "").replace("_", "").replace(".", "").replace(" ", "")
        if suffix in self.files_by_suffix:
            for file, normalized_name in self.files_by_suffix[suffix]:
                if normalized_name.startswith(normalized_code) and normalized_name.endswith(suffix.lower()):
                    return file
        return None

    def pause(self):
        self._pause_event.clear()
        self._log(None, "⏸️ Đã tạm dừng", 'info')

    def resume(self):
        self._pause_event.set()
        self._log(None, "▶️ Đã tiếp tục", 'info')

    def stop(self):
        self._stop_event.set()
        self._pause_event.set()
        self._log(None, "🛑 Đã dừng", 'info')